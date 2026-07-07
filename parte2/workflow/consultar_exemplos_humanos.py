from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "fonte_parte_2.xlsx"
SHEET = "municipios_classificados"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def main():
    terms = [term.upper() for term in sys.argv[1:]]
    if not terms:
        raise SystemExit("Informe termos de busca.")

    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb[SHEET]
    headers = [clean(cell.value) for cell in ws[1]]
    area_header = next((header for header in headers if "Tem" in header and "tica" in header), "Area Tematica")

    matches = {term: [] for term in terms}
    fields = ["nomePrograma", "nomeFuncao", "nomeSubFuncao", "nomeAcaoOrcamentaria"]

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = {headers[i]: clean(value) for i, value in enumerate(row) if i < len(headers)}
        haystack = " | ".join(record.get(field, "") for field in fields).upper()
        for term in terms:
            if term in haystack and len(matches[term]) < 8:
                matches[term].append(
                    {
                        "excel_row": row_number,
                        "nomeMunicipio": record.get("nomeMunicipio", ""),
                        "nomePrograma": record.get("nomePrograma", ""),
                        "nomeFuncao": record.get("nomeFuncao", ""),
                        "nomeSubFuncao": record.get("nomeSubFuncao", ""),
                        "nomeAcaoOrcamentaria": record.get("nomeAcaoOrcamentaria", ""),
                        "Area Tematica": record.get(area_header, ""),
                        "Gasto E ou NE": record.get("Gasto E ou NE", ""),
                        "Indicador": record.get("Indicador", ""),
                    }
                )

    print(json.dumps(matches, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
