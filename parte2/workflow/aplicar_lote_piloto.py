from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
WORKING_COPY = ROOT / "parte2" / "resultados" / "fonte_parte_2_trabalho.xlsx"
CLASSIFIED_CSV = ROOT / "parte2" / "resultados" / "lote_piloto_001_classificado.csv"
SHEET = "classificacao_automatizada"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def main():
    with CLASSIFIED_CSV.open("r", newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))

    wb = load_workbook(WORKING_COPY)
    ws = wb[SHEET]
    headers = [clean(cell.value) for cell in ws[1]]

    area_col = next(
        index + 1
        for index, header in enumerate(headers)
        if header in {"Area Tematica", "Área Temática", "Ãrea TemÃ¡tica"}
    )
    gasto_col = headers.index("Gasto E ou NE") + 1

    for row in rows:
        excel_row = int(row["_excel_row"])
        ws.cell(row=excel_row, column=area_col).value = row["Area Tematica LLM"]
        ws.cell(row=excel_row, column=gasto_col).value = row["Gasto E ou NE LLM"]

    wb.save(WORKING_COPY)
    print(f"applied={len(rows)} workbook={WORKING_COPY.relative_to(ROOT)} sheet={SHEET}")


if __name__ == "__main__":
    main()
