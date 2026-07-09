from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "parte2" / "resultados" / "fonte_parte_2_trabalho.xlsx"
RESULTS = ROOT / "parte2" / "resultados"
SHEET = "classificacao_automatizada"

ALIASES = {
    "Area Tematica": ["Area Tematica", "Ãrea TemÃ¡tica", "ÃƒÂrea TemÃƒÂ¡tica"],
    "nomeMunicipio": ["nomeMunicipio", "municipio", "Municipio"],
}


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def find_header(headers, canonical):
    for candidate in ALIASES.get(canonical, [canonical]):
        if candidate in headers:
            return candidate
    if canonical == "Area Tematica":
        for header in headers:
            normalized = header.lower()
            if "rea" in normalized and "tem" in normalized and "tica" in normalized:
                return header
    return canonical


def write_csv(path, rows, fields):
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main():
    if len(sys.argv) != 4:
        raise SystemExit("Uso: python preparar_lote.py <lote_id> <excel_row_inicio> <quantidade>")

    lote_id = sys.argv[1]
    start_row = int(sys.argv[2])
    size = int(sys.argv[3])
    end_row = start_row + size - 1

    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb[SHEET]
    headers = [clean(cell.value) for cell in ws[1]]
    area_header = find_header(headers, "Area Tematica")
    municipio_header = find_header(headers, "nomeMunicipio")

    rows = []
    for row_number, row in enumerate(ws.iter_rows(min_row=start_row, max_row=end_row, values_only=True), start=start_row):
        record = {headers[i]: clean(value) for i, value in enumerate(row) if i < len(headers)}
        record["_excel_row"] = row_number
        record["Area Tematica"] = record.get(area_header, "")
        record["nomeMunicipio"] = record.get(municipio_header, "")
        rows.append(record)

    fields = [
        "_excel_row",
        "Id",
        "nomeMunicipio",
        "nomePrograma",
        "nomeFuncao",
        "nomeSubFuncao",
        "nomeAcaoOrcamentaria",
        "Area Tematica",
        "Gasto E ou NE",
    ]
    fields = [field for field in fields if field in {"_excel_row", "Area Tematica", "nomeMunicipio"} or field in headers]
    output = RESULTS / f"{lote_id}.csv"
    write_csv(output, rows, fields)
    print(f"lote={lote_id} rows={len(rows)} excel_rows={start_row}-{end_row} output={output.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
