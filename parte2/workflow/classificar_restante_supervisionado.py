from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
RESULTS = ROOT / "parte2" / "resultados"
WORKBOOK = RESULTS / "fonte_parte_2_trabalho.xlsx"
SHEET = "classificacao_automatizada"
LOG_CSV = RESULTS / "classificacao_automatica_supervisionada_log.csv"
SUMMARY_JSON = RESULTS / "classificacao_automatica_supervisionada_resumo.json"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def norm(value):
    text = clean(value).lower()
    table = str.maketrans(
        "áàãâäéèêëíìîïóòõôöúùûüçñ",
        "aaaaaeeeeiiiiooooouuuucn",
    )
    text = text.translate(table)
    return re.sub(r"\s+", " ", text)


def has(text, *terms):
    return any(term in text for term in terms)


def classify(record):
    programa = norm(record.get("nomePrograma"))
    funcao = norm(record.get("nomeFuncao"))
    subfuncao = norm(record.get("nomeSubFuncao"))
    acao = norm(record.get("nomeAcaoOrcamentaria"))
    text = " | ".join([programa, funcao, subfuncao, acao])

    def decision(area, gasto, conf, reason, review=False, doubts=""):
        return {
            "Area Tematica LLM": area,
            "Gasto E ou NE LLM": gasto,
            "Confianca": conf,
            "Justificativa": reason,
            "Campos faltantes ou duvidas": doubts,
            "Requer revisao humana": "Sim" if review else "Nao",
        }

    if has(text, "ensino infantil", "educacao infantil", "creche", "pre escola", "pre-escola") or (
        "primeira infancia" in programa and funcao == "educacao"
    ):
        return decision(
            "Educacao Infantil",
            "Especifico",
            "Alta",
            "A linha explicita ensino infantil, educacao infantil, creche/pre-escola ou programa de primeira infancia em educacao.",
        )

    if "alimentacao escolar" in acao or "merenda escolar" in acao:
        if has(text, "infantil", "creche", "pre escola", "pre-escola"):
            return decision("Educacao Infantil", "Especifico", "Alta", "Alimentacao/merenda escolar explicitamente vinculada a educacao infantil.")
        if has(text, "fundamental", "eja", "ensino medio"):
            return decision("nao se aplica", "-", "Alta", "Alimentacao escolar vinculada a etapa fora da educacao infantil.")
        return decision(
            "Educacao Infantil",
            "Nao Especifico",
            "Media",
            "Alimentacao escolar pode abranger educacao infantil, mas a etapa nao esta explicita.",
            True,
            "Confirmar se a alimentacao escolar abrange educacao infantil.",
        )

    if "transporte escolar" in acao:
        if has(text, "infantil", "creche", "pre escola", "pre-escola"):
            return decision("Educacao Infantil", "Especifico", "Alta", "Transporte escolar explicitamente vinculado a educacao infantil.")
        if "fundamental" in text:
            return decision("nao se aplica", "-", "Alta", "Transporte escolar vinculado ao ensino fundamental.")
        return decision(
            "Educacao Infantil",
            "Nao Especifico",
            "Media",
            "Transporte escolar pode abranger educacao infantil, mas a etapa nao esta explicita.",
            True,
            "Confirmar se o transporte escolar abrange educacao infantil.",
        )

    if funcao == "educacao" and ("fundamental" in text or "ensino superior" in text or "ensino medio" in text or "eja" in text):
        return decision("nao se aplica", "-", "Alta", "Acao vinculada a etapa educacional fora da educacao infantil.")

    if funcao == "educacao" and has(acao, "fundo municipal de educacao", " fme", "- fme"):
        if "administracao geral" in subfuncao or "gestao" in programa:
            return decision(
                "Educacao Infantil",
                "Nao Especifico",
                "Media",
                "Despesa-meio do FME pode sustentar politicas de educacao infantil, mas a etapa nao esta explicita.",
                True,
                "Confirmar manutencao do padrao FME em administracao geral.",
            )

    if funcao == "educacao" and has(text, "apoio pedagogico", "material didatico", "reaparelhamento", "unidade escolar", "educacao em tempo integral"):
        return decision(
            "Educacao Infantil",
            "Nao Especifico",
            "Media",
            "Acao educacional ampla pode beneficiar educacao infantil, mas nao explicita exclusividade da etapa.",
            True,
            "Confirmar abrangencia da educacao infantil.",
        )

    if funcao == "saude" or has(text, "fms", "saude", "ubs", "atencao basica", "vigilancia sanitaria", "vigilancia epidemiologica", "saude bucal", "assistencia farmaceutica", "agente comunitario", "estrategia saude da familia", "conselho municipal de saude"):
        if has(text, "inativo", "aposentado", "previdencia"):
            return decision("nao se aplica", "-", "Alta", "Despesa de inativos/previdencia fora do escopo.")
        return decision(
            "Saude Materno-infantil",
            "Nao Especifico",
            "Alta",
            "Acao de saude ou despesa-meio de saude beneficia criancas pequenas, gestantes e familias de forma ampliada.",
        )

    if has(text, "primeira infancia") and funcao == "assistencia social":
        return decision("Assistencia Social", "Especifico", "Alta", "Acao de assistencia social explicitamente vinculada a primeira infancia.")

    if funcao == "assistencia social" or has(text, "cras", "creas", "fmas", "suas", "beneficio eventual", "assistencia social", "bolsa familia", "cadunico", "cadastro unico", "crianca feliz"):
        return decision(
            "Assistencia Social",
            "Nao Especifico",
            "Alta",
            "Acao de assistencia social atende familias e individuos, incluindo familias com criancas pequenas de forma ampliada.",
        )

    if has(text, "conselho tutelar", "crianca e ao adolescente", "crianca e adolescente"):
        return decision(
            "Protecao dos Direitos da Crianca e da Familia",
            "Nao Especifico",
            "Media",
            "Protecao de direitos de criancas e adolescentes inclui primeira infancia, mas nao e exclusiva de 0 a 6 anos.",
            True,
            "Confirmar enquadramento em protecao de direitos como gasto ampliado.",
        )

    if has(text, "politicas para mulheres", "direitos das mulheres", "mulheres"):
        return decision(
            "Protecao dos Direitos da Crianca e da Familia",
            "Nao Especifico",
            "Media",
            "Politicas para mulheres podem beneficiar gestantes, lactantes e familias de forma ampliada.",
            True,
            "Confirmar se ha acoes com gestantes, lactantes ou familias com primeira infancia.",
        )

    if has(text, "saneamento", "agua", "esgoto", "aterro sanitario", "residuos", "limpeza publica", "coleta de lixo", "drenagem"):
        if has(text, "iluminacao publica", "estrada", "via urbana", "pavimentacao"):
            return decision("nao se aplica", "-", "Alta", "Infraestrutura urbana genericamente excluida sem beneficio concreto ao publico da primeira infancia.")
        return decision(
            "Saneamento e Agua",
            "Nao Especifico",
            "Media" if funcao == "gestao ambiental" else "Alta",
            "Acao relacionada a saneamento, agua, limpeza urbana ou residuos beneficia a populacao de forma ampliada.",
            funcao == "gestao ambiental",
            "Confirmar enquadramento quando a funcao for gestao ambiental." if funcao == "gestao ambiental" else "",
        )

    if has(text, "seguranca alimentar", "cesta basica", "alimentos", "nutricao") and funcao != "educacao":
        return decision(
            "Seguranca Alimentar",
            "Nao Especifico",
            "Media",
            "Acao de alimentacao ou seguranca alimentar pode beneficiar familias com primeira infancia de forma ampliada.",
            True,
            "Confirmar publico-alvo e abrangencia.",
        )

    if has(text, "habitacao", "moradia", "habitacional", "regularizacao fundiaria", "praca", "parque", "playground"):
        if has(text, "crianca", "infantil", "primeira infancia", "playground"):
            return decision(
                "Direito a Cidade e Habitacao",
                "Nao Especifico",
                "Media",
                "Acao urbana/habitacional com indicio de beneficio a criancas ou familias.",
                True,
                "Confirmar beneficio concreto ao publico de primeira infancia.",
            )

    if has(text, "cultura", "brincar", "lazer", "esporte", "desporto"):
        if has(text, "crianca", "infantil", "primeira infancia", "brincar"):
            return decision(
                "Cultura e Direito de Brincar",
                "Nao Especifico",
                "Media",
                "Acao de cultura, lazer ou brincar com indicio de publico infantil.",
                True,
                "Confirmar foco em criancas de 0 a 6 anos.",
            )
        return decision("nao se aplica", "-", "Alta", "Cultura, esporte ou lazer generico sem publico da primeira infancia explicitado.")

    if has(text, "transferencia de renda", "auxilio brasil", "renda", "enfrentamento da pobreza"):
        return decision(
            "Enfrentamento da Pobreza",
            "Nao Especifico",
            "Media",
            "Acao de renda ou enfrentamento da pobreza pode beneficiar familias com primeira infancia de forma ampliada.",
            True,
            "Confirmar publico-alvo e criterio do beneficio.",
        )

    if has(text, "legislativa", "camara municipal", "gabinete", "administracao", "financas", "planejamento", "comunicacao", "tecnologia", "previdencia", "precatorio", "sentenca judicial", "agricultura", "pecuaria", "estrada vicinal", "iluminacao publica", "transporte rodoviario", "meio ambiente", "turismo"):
        return decision("nao se aplica", "-", "Alta", "Acao administrativa, legislativa, setorial generica ou infraestrutura excluida sem beneficio concreto ao GSPI-M.")

    return decision(
        "nao se aplica",
        "-",
        "Media",
        "Nao foram encontrados indicios suficientes de conexao com as areas finalisticas do GSPI-M.",
        True,
        "Revisar por ausencia de padrao especifico.",
    )


def main():
    wb = load_workbook(WORKBOOK)
    ws = wb[SHEET]
    headers = [clean(cell.value) for cell in ws[1]]
    col = {header: index + 1 for index, header in enumerate(headers)}
    area_col = next(i + 1 for i, header in enumerate(headers) if "rea" in header.lower() and "tem" in header.lower())
    gasto_col = col["Gasto E ou NE"]
    indicador_col = col["Indicador"]

    fields = [
        "excel_row",
        "Id",
        "nomeMunicipio",
        "nomePrograma",
        "nomeFuncao",
        "nomeSubFuncao",
        "nomeAcaoOrcamentaria",
        "Area Tematica LLM",
        "Gasto E ou NE LLM",
        "Confianca",
        "Justificativa",
        "Campos faltantes ou duvidas",
        "Requer revisao humana",
    ]

    log_rows = []
    before = 0
    applied = 0
    skipped = 0
    touched_indicator = 0

    for row_num in range(2, ws.max_row + 1):
        current_area = clean(ws.cell(row_num, area_col).value)
        current_gasto = clean(ws.cell(row_num, gasto_col).value)
        if current_area or current_gasto:
            before += 1
            continue

        record = {header: clean(ws.cell(row_num, idx).value) for header, idx in col.items()}
        result = classify(record)
        ws.cell(row_num, area_col).value = result["Area Tematica LLM"]
        ws.cell(row_num, gasto_col).value = result["Gasto E ou NE LLM"]
        if clean(ws.cell(row_num, indicador_col).value):
            touched_indicator += 1

        log_rows.append(
            {
                "excel_row": row_num,
                "Id": record.get("Id", ""),
                "nomeMunicipio": record.get("nomeMunicipio", ""),
                "nomePrograma": record.get("nomePrograma", ""),
                "nomeFuncao": record.get("nomeFuncao", ""),
                "nomeSubFuncao": record.get("nomeSubFuncao", ""),
                "nomeAcaoOrcamentaria": record.get("nomeAcaoOrcamentaria", ""),
                **result,
            }
        )
        applied += 1

    with LOG_CSV.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(log_rows)

    wb.save(WORKBOOK)

    area_counts = Counter(row["Area Tematica LLM"] for row in log_rows)
    gasto_counts = Counter(row["Gasto E ou NE LLM"] for row in log_rows)
    confianca_counts = Counter(row["Confianca"] for row in log_rows)
    municipio_counts = Counter(row["nomeMunicipio"] for row in log_rows)
    revisao_counts = Counter(row["Requer revisao humana"] for row in log_rows)

    summary = {
        "workbook": str(WORKBOOK.relative_to(ROOT)),
        "sheet": SHEET,
        "linhas_ja_preenchidas_antes": before,
        "linhas_aplicadas": applied,
        "linhas_ignoradas": skipped,
        "indicador_nao_vazio_nas_linhas_aplicadas": touched_indicator,
        "distribuicao_area": dict(area_counts),
        "distribuicao_gasto": dict(gasto_counts),
        "distribuicao_confianca": dict(confianca_counts),
        "distribuicao_revisao": dict(revisao_counts),
        "distribuicao_municipio": dict(municipio_counts),
        "log": str(LOG_CSV.relative_to(ROOT)),
    }
    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()
