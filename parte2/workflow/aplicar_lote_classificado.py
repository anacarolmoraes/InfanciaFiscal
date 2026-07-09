from __future__ import annotations

import csv
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
WORKING_COPY = ROOT / "parte2" / "resultados" / "fonte_parte_2_trabalho.xlsx"
SHEET = "classificacao_automatizada"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Uso: python aplicar_lote_classificado.py <csv_classificado>")

    classified_csv = ROOT / sys.argv[1]
    with classified_csv.open("r", newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))

    wb = load_workbook(WORKING_COPY)
    ws = wb[SHEET]
    headers = [clean(cell.value) for cell in ws[1]]

    area_col = next(
        index + 1
        for index, header in enumerate(headers)
        if header in {"Area Tematica", "Ãrea TemÃ¡tica", "ÃƒÂrea TemÃƒÂ¡tica"}
        or ("rea" in header.lower() and "tem" in header.lower() and "tica" in header.lower())
    )
    gasto_col = headers.index("Gasto E ou NE") + 1

    for row in rows:
        excel_row = int(row["_excel_row"])
        ws.cell(row=excel_row, column=area_col).value = row["Area Tematica LLM"]
        ws.cell(row=excel_row, column=gasto_col).value = row["Gasto E ou NE LLM"]

    wb.save(WORKING_COPY)
    print(f"applied={len(rows)} workbook={WORKING_COPY.relative_to(ROOT)} sheet={SHEET}")


if __name__ == "__main__":
    main()
