from __future__ import annotations

import csv
import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "fonte_parte_2.xlsx"
RESULTS = ROOT / "parte2" / "resultados"

BASE_SHEET = "municipios_classificados"
TARGET_SHEET = "classificacao_automatizada"

WORKING_COPY = RESULTS / "fonte_parte_2_trabalho.xlsx"
SUMMARY_JSON = RESULTS / "preparacao_execucao_resumo.json"
EXAMPLES_CSV = RESULTS / "exemplos_humanos_representativos.csv"
PILOT_CSV = RESULTS / "lote_piloto_001.csv"
LOG_TEMPLATE_CSV = RESULTS / "log_classificacao_template.csv"
NORMALIZATION_JSON = RESULTS / "normalizacao_labels_inicial.json"

REQUIRED_FIELDS = [
    "nomePrograma",
    "nomeFuncao",
    "nomeSubFuncao",
    "nomeAcaoOrcamentaria",
]

OUTPUT_FIELDS = [
    "Area Tematica",
    "Gasto E ou NE",
]

ALIASES = {
    "Area Tematica": ["Area Tematica", "Área Temática", "Ãrea TemÃ¡tica"],
    "Gasto E ou NE": ["Gasto E ou NE"],
    "nomeMunicipio": ["nomeMunicipio", "municipio", "Municipio"],
}


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def sheet_rows(ws):
    headers = [clean(cell.value) for cell in ws[1]]
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = {headers[i]: clean(value) for i, value in enumerate(row) if i < len(headers)}
        record["_excel_row"] = row_number
        yield record


def find_header(headers, canonical):
    for candidate in ALIASES.get(canonical, [canonical]):
        if candidate in headers:
            return candidate
    return ""


def canonicalize_rows(rows, headers):
    area_field = find_header(headers, "Area Tematica")
    municipio_field = find_header(headers, "nomeMunicipio")
    output = []
    for row in rows:
        item = dict(row)
        if area_field:
            item["Area Tematica"] = row.get(area_field, "")
        if municipio_field:
            item["nomeMunicipio"] = row.get(municipio_field, "")
        output.append(item)
    return output


def write_csv(path, rows, fields):
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main():
    RESULTS.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    if BASE_SHEET not in wb.sheetnames or TARGET_SHEET not in wb.sheetnames:
        raise SystemExit(f"Abas esperadas ausentes: {BASE_SHEET}, {TARGET_SHEET}")

    base_ws = wb[BASE_SHEET]
    target_ws = wb[TARGET_SHEET]

    base_headers = [clean(cell.value) for cell in base_ws[1]]
    target_headers = [clean(cell.value) for cell in target_ws[1]]

    base_rows = canonicalize_rows(list(sheet_rows(base_ws)), base_headers)
    target_rows = canonicalize_rows(list(sheet_rows(target_ws)), target_headers)

    area_field = "Area Tematica"
    gasto_field = "Gasto E ou NE"

    area_counts = Counter(row.get(area_field, "") for row in base_rows)
    gasto_counts = Counter(row.get(gasto_field, "") for row in base_rows)
    combo_counts = Counter((row.get(area_field, ""), row.get(gasto_field, "")) for row in base_rows)

    examples_by_combo = defaultdict(list)
    for row in base_rows:
        combo = (row.get(area_field, ""), row.get(gasto_field, ""))
        if len(examples_by_combo[combo]) < 2:
            examples_by_combo[combo].append(row)

    examples = []
    for combo in sorted(examples_by_combo):
        examples.extend(examples_by_combo[combo])

    example_fields = [
        "_excel_row",
        "Id",
        "nomeMunicipio",
        "nomePrograma",
        "nomeFuncao",
        "nomeSubFuncao",
        "nomeAcaoOrcamentaria",
        "Area Tematica",
        "Gasto E ou NE",
        "Indicador",
    ]
    example_fields = [
        field
        for field in example_fields
        if field in {"_excel_row", "Area Tematica", "nomeMunicipio"} or field in base_headers
    ]
    write_csv(EXAMPLES_CSV, examples, example_fields)

    pilot_rows = target_rows[:20]
    pilot_fields = [
        "_excel_row",
        "Id",
        "nomeMunicipio",
        "nomePrograma",
        "nomeFuncao",
        "nomeSubFuncao",
        "nomeAcaoOrcamentaria",
        "Area Tematica",
        "Gasto E ou NE",
    ]
    pilot_fields = [
        field
        for field in pilot_fields
        if field in {"_excel_row", "Area Tematica", "nomeMunicipio"} or field in target_headers
    ]
    write_csv(PILOT_CSV, pilot_rows, pilot_fields)

    log_fields = [
        "lote",
        "excel_row",
        "Id",
        "nomeMunicipio",
        "nomePrograma",
        "nomeFuncao",
        "nomeSubFuncao",
        "nomeAcaoOrcamentaria",
        "Area Tematica LLM",
        "Gasto E ou NE LLM",
        "Confianca",
        "Justificativa",
        "Campos faltantes ou duvidas",
        "Requer revisao humana",
    ]
    write_csv(LOG_TEMPLATE_CSV, [], log_fields)

    summary = {
        "arquivo_operacional": str(WORKBOOK.relative_to(ROOT)),
        "copia_trabalho": str(WORKING_COPY.relative_to(ROOT)),
        "abas": wb.sheetnames,
        "municipios_classificados": {
            "linhas_dados": len(base_rows),
            "colunas": len(base_headers),
            "headers": base_headers,
            "header_area_tematica_detectado": find_header(base_headers, "Area Tematica"),
            "campos_requeridos_presentes": {field: field in base_headers for field in REQUIRED_FIELDS},
            "campos_saida_presentes": {
                "Area Tematica": bool(find_header(base_headers, "Area Tematica")),
                "Gasto E ou NE": "Gasto E ou NE" in base_headers,
            },
            "area_tematica_contagem": dict(area_counts),
            "gasto_e_ou_ne_contagem": dict(gasto_counts),
            "combinacoes_area_gasto": {f"{area} | {gasto}": count for (area, gasto), count in combo_counts.items()},
        },
        "classificacao_automatizada": {
            "linhas_dados": len(target_rows),
            "colunas": len(target_headers),
            "headers": target_headers,
            "header_area_tematica_detectado": find_header(target_headers, "Area Tematica"),
            "campos_requeridos_presentes": {field: field in target_headers for field in REQUIRED_FIELDS},
            "campos_saida_presentes": {
                "Area Tematica": bool(find_header(target_headers, "Area Tematica")),
                "Gasto E ou NE": "Gasto E ou NE" in target_headers,
            },
            "lote_piloto": {
                "arquivo": str(PILOT_CSV.relative_to(ROOT)),
                "linhas": len(pilot_rows),
                "excel_rows": [row["_excel_row"] for row in pilot_rows],
            },
            "tamanho_lote_recomendado": 20,
        },
        "artefatos_gerados": [
            str(SUMMARY_JSON.relative_to(ROOT)),
            str(EXAMPLES_CSV.relative_to(ROOT)),
            str(PILOT_CSV.relative_to(ROOT)),
            str(LOG_TEMPLATE_CSV.relative_to(ROOT)),
            str(NORMALIZATION_JSON.relative_to(ROOT)),
        ],
    }

    normalization = {
        "observacao": "Mapa inicial para avaliacao; nao altera a planilha original.",
        "area_tematica_valores_observados": sorted(area_counts),
        "gasto_e_ou_ne_valores_observados": sorted(gasto_counts),
        "normalizacoes_sugeridas": {
            "Nao especifico": "Nao Especifico",
            "Assitencia Social": "Assistencia Social",
            "Sanemaneto e agua": "Saneamento e Agua",
            "-": "",
            "nao se aplica": "",
        },
    }

    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    NORMALIZATION_JSON.write_text(json.dumps(normalization, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps({
        "status": "ok",
        "base_rows": len(base_rows),
        "target_rows": len(target_rows),
        "pilot_rows": len(pilot_rows),
        "examples_rows": len(examples),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
