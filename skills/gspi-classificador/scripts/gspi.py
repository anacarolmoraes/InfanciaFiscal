from __future__ import annotations

import argparse
import csv
import json
import random
import re
import shutil
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_WORKBOOK = Path("parte2/resultados/fonte_parte_2_trabalho.xlsx")
DEFAULT_SHEET = "classificacao_automatizada"
DEFAULT_LOG = Path("parte2/resultados/classificacao_automatica_supervisionada_log.xlsx")
DEFAULT_SUMMARY = Path("parte2/resultados/classificacao_automatica_supervisionada_resumo.json")


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def norm(value):
    text = unicodedata.normalize("NFKD", clean(value).lower())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"\s+", " ", text)


def has(text, *terms):
    return any(term in text for term in terms)


def headers_for(ws):
    return [clean(cell.value) for cell in ws[1]]


def read_csv_rows(path):
    for encoding in ("utf-8-sig", "cp1252", "latin1"):
        try:
            with open(path, newline="", encoding=encoding) as fh:
                return list(csv.DictReader(fh))
        except UnicodeDecodeError:
            continue
    raise SystemExit(f"Nao foi possivel ler o CSV: {path}")


def read_table_rows(path):
    path = Path(path)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [clean(value) for value in rows[0]]
        return [
            {headers[index]: clean(value) for index, value in enumerate(row) if index < len(headers)}
            for row in rows[1:]
        ]
    return read_csv_rows(path)


def write_table_rows(path, fields, rows, sheet_name="dados"):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(fields)
        for row in rows:
            ws.append([row.get(field, "") for field in fields])
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in ws[1]:
            cell.font = Font(name="Arial", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial")
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column_cells in ws.columns:
            values = [clean(cell.value) for cell in column_cells]
            width = min(max([len(value) for value in values] + [12]) + 2, 60)
            ws.column_dimensions[column_cells[0].column_letter].width = width
        ws.freeze_panes = "A2"
        wb.save(path)
        return
    with open(path, "w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def find_area_col(headers):
    for index, header in enumerate(headers):
        h = norm(header)
        if "area" in h and "tematica" in h:
            return index + 1
    raise SystemExit("Coluna Area Tematica nao encontrada.")


def classify(record):
    programa = norm(record.get("nomePrograma"))
    funcao = norm(record.get("nomeFuncao"))
    acao = norm(record.get("nomeAcaoOrcamentaria"))
    text = " | ".join(
        [
            programa,
            funcao,
            norm(record.get("nomeSubFuncao")),
            acao,
        ]
    )

    def decision(area, gasto, conf, reason, review=False, doubts=""):
        return {
            "Area Tematica LLM": area,
            "Gasto E ou NE LLM": gasto,
            "Confianca": conf,
            "Justificativa": reason,
            "Campos faltantes ou duvidas": doubts,
            "Requer revisao humana": "Sim" if review else "Nao",
        }

    if has(text, "ensino infantil", "educacao infantil", "creche", "pre escola", "pre-escola") or (
        "primeira infancia" in programa and funcao == "educacao"
    ):
        return decision("Educacao Infantil", "Especifico", "Alta", "Referencia explicita a educacao infantil, creche, pre-escola ou primeira infancia em educacao.")

    if "alimentacao escolar" in acao or "merenda escolar" in acao:
        if has(text, "infantil", "creche", "pre escola", "pre-escola"):
            return decision("Educacao Infantil", "Especifico", "Alta", "Alimentacao escolar vinculada explicitamente a educacao infantil.")
        if has(text, "fundamental", "eja", "ensino medio"):
            return decision("nao se aplica", "-", "Alta", "Alimentacao escolar vinculada a etapa fora da educacao infantil.")
        return decision("Educacao Infantil", "Nao Especifico", "Media", "Alimentacao escolar pode abranger educacao infantil, mas a etapa nao esta explicita.", True, "Confirmar abrangencia da educacao infantil.")

    if "transporte escolar" in acao:
        if has(text, "infantil", "creche", "pre escola", "pre-escola"):
            return decision("Educacao Infantil", "Especifico", "Alta", "Transporte escolar vinculado explicitamente a educacao infantil.")
        if "fundamental" in text:
            return decision("nao se aplica", "-", "Alta", "Transporte escolar vinculado ao ensino fundamental.")
        return decision("Educacao Infantil", "Nao Especifico", "Media", "Transporte escolar pode abranger educacao infantil, mas a etapa nao esta explicita.", True, "Confirmar abrangencia da educacao infantil.")

    if funcao == "educacao" and has(text, "fundamental", "ensino superior", "ensino medio", "eja"):
        return decision("nao se aplica", "-", "Alta", "Acao vinculada a etapa educacional fora da educacao infantil.")

    if funcao == "educacao" and has(text, "fundo municipal de educacao", " fme", "- fme"):
        return decision("Educacao Infantil", "Nao Especifico", "Media", "Despesa-meio do FME pode sustentar educacao infantil, mas a etapa nao esta explicita.", True, "Confirmar padrao FME.")

    if funcao == "educacao" and has(text, "apoio pedagogico", "material didatico", "reaparelhamento", "unidade escolar", "educacao em tempo integral"):
        return decision("Educacao Infantil", "Nao Especifico", "Media", "Acao educacional ampla pode beneficiar educacao infantil, mas nao explicita exclusividade.", True, "Confirmar abrangencia.")

    if funcao == "saude" or has(text, "fms", "saude", "ubs", "atencao basica", "vigilancia sanitaria", "vigilancia epidemiologica", "saude bucal", "assistencia farmaceutica", "agente comunitario", "estrategia saude da familia", "conselho municipal de saude"):
        if has(text, "inativo", "aposentado", "previdencia"):
            return decision("nao se aplica", "-", "Alta", "Despesa de inativos ou previdencia fora do escopo.")
        return decision("Saude Materno-infantil", "Nao Especifico", "Alta", "Acao ou despesa-meio de saude beneficia primeira infancia, gestantes e familias de forma ampliada.")

    if has(text, "primeira infancia") and funcao == "assistencia social":
        return decision("Assistencia Social", "Especifico", "Alta", "Assistencia social explicitamente vinculada a primeira infancia.")

    if funcao == "assistencia social" or has(text, "cras", "creas", "fmas", "suas", "beneficio eventual", "assistencia social", "bolsa familia", "cadunico", "cadastro unico", "crianca feliz"):
        return decision("Assistencia Social", "Nao Especifico", "Alta", "Acao de assistencia social atende familias e individuos, incluindo primeira infancia de forma ampliada.")

    if has(text, "conselho tutelar", "crianca e ao adolescente", "crianca e adolescente"):
        return decision("Protecao dos Direitos da Crianca e da Familia", "Nao Especifico", "Media", "Protecao de direitos de criancas e adolescentes inclui primeira infancia, mas nao e exclusiva de 0 a 6 anos.", True, "Confirmar enquadramento.")

    if has(text, "politicas para mulheres", "direitos das mulheres", "mulheres"):
        return decision("Protecao dos Direitos da Crianca e da Familia", "Nao Especifico", "Media", "Politicas para mulheres podem beneficiar gestantes, lactantes e familias de forma ampliada.", True, "Confirmar publico-alvo.")

    if has(text, "saneamento", "agua", "esgoto", "aterro sanitario", "residuos", "limpeza publica", "coleta de lixo", "drenagem"):
        if has(text, "iluminacao publica", "estrada", "pavimentacao"):
            return decision("nao se aplica", "-", "Alta", "Infraestrutura urbana generica excluida sem beneficio concreto.")
        review = funcao == "gestao ambiental"
        return decision("Saneamento e Agua", "Nao Especifico", "Media" if review else "Alta", "Saneamento, agua, limpeza urbana ou residuos beneficiam a populacao de forma ampliada.", review, "Confirmar quando estiver em gestao ambiental." if review else "")

    if has(text, "seguranca alimentar", "cesta basica", "alimentos", "nutricao") and funcao != "educacao":
        return decision("Seguranca Alimentar", "Nao Especifico", "Media", "Acao de alimentacao ou seguranca alimentar pode beneficiar familias com primeira infancia.", True, "Confirmar publico-alvo.")

    if has(text, "habitacao", "moradia", "habitacional", "regularizacao fundiaria", "praca", "parque", "playground"):
        if has(text, "crianca", "infantil", "primeira infancia", "playground"):
            return decision("Direito a Cidade e Habitacao", "Nao Especifico", "Media", "Acao urbana ou habitacional com indicio de beneficio a criancas ou familias.", True, "Confirmar beneficio concreto.")

    if has(text, "cultura", "brincar", "lazer", "esporte", "desporto"):
        if has(text, "crianca", "infantil", "primeira infancia", "brincar"):
            return decision("Cultura e Direito de Brincar", "Nao Especifico", "Media", "Cultura, lazer ou brincar com indicio de publico infantil.", True, "Confirmar foco em 0 a 6 anos.")
        return decision("nao se aplica", "-", "Alta", "Cultura, esporte ou lazer generico sem primeira infancia explicitada.")

    if has(text, "transferencia de renda", "auxilio brasil", "renda", "enfrentamento da pobreza"):
        return decision("Enfrentamento da Pobreza", "Nao Especifico", "Media", "Acao de renda ou enfrentamento da pobreza pode beneficiar familias com primeira infancia.", True, "Confirmar publico-alvo.")

    if has(text, "legislativa", "camara municipal", "gabinete", "administracao", "financas", "planejamento", "comunicacao", "tecnologia", "previdencia", "precatorio", "sentenca judicial", "agricultura", "pecuaria", "estrada vicinal", "iluminacao publica", "transporte rodoviario", "meio ambiente", "turismo"):
        return decision("nao se aplica", "-", "Alta", "Acao administrativa, legislativa, setorial generica ou infraestrutura excluida sem beneficio concreto ao GSPI-M.")

    return decision("nao se aplica", "-", "Media", "Sem indicios suficientes de conexao com as areas finalisticas do GSPI-M.", True, "Revisar por ausencia de padrao especifico.")


def command_prepare(args):
    source = Path(args.source)
    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.exists() and not args.force:
        raise SystemExit(f"Arquivo ja existe: {output}. Use --force para sobrescrever.")
    shutil.copy2(source, output)
    print(json.dumps({"status": "ok", "source": str(source), "output": str(output)}, ensure_ascii=False))


def workbook_status(workbook, sheet):
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb[sheet]
    headers = headers_for(ws)
    area_i = find_area_col(headers) - 1
    gasto_i = headers.index("Gasto E ou NE")
    indicador_i = headers.index("Indicador")
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return {
        "rows": len(rows),
        "area_filled": sum(1 for row in rows if row[area_i] not in (None, "")),
        "gasto_filled": sum(1 for row in rows if row[gasto_i] not in (None, "")),
        "indicador_filled": sum(1 for row in rows if row[indicador_i] not in (None, "")),
    }


def command_status(args):
    print(json.dumps(workbook_status(args.workbook, args.sheet), ensure_ascii=False))


def command_classify(args):
    wb = load_workbook(args.workbook)
    ws = wb[args.sheet]
    headers = headers_for(ws)
    col = {header: index + 1 for index, header in enumerate(headers)}
    area_col = find_area_col(headers)
    gasto_col = col["Gasto E ou NE"]
    indicador_col = col["Indicador"]
    fields = ["excel_row", "Id", "nomeMunicipio", "nomePrograma", "nomeFuncao", "nomeSubFuncao", "nomeAcaoOrcamentaria", "Area Tematica LLM", "Gasto E ou NE LLM", "Confianca", "Justificativa", "Campos faltantes ou duvidas", "Requer revisao humana"]
    rows = []
    before = 0
    indicator_nonempty_on_applied_rows = 0

    for row_num in range(2, ws.max_row + 1):
        if clean(ws.cell(row_num, area_col).value) or clean(ws.cell(row_num, gasto_col).value):
            before += 1
            continue
        record = {header: clean(ws.cell(row_num, idx).value) for header, idx in col.items()}
        result = classify(record)
        ws.cell(row_num, area_col).value = result["Area Tematica LLM"]
        ws.cell(row_num, gasto_col).value = result["Gasto E ou NE LLM"]
        if clean(ws.cell(row_num, indicador_col).value):
            indicator_nonempty_on_applied_rows += 1
        rows.append(
            {
                "excel_row": row_num,
                "Id": record.get("Id", ""),
                "nomeMunicipio": record.get("nomeMunicipio", ""),
                "nomePrograma": record.get("nomePrograma", ""),
                "nomeFuncao": record.get("nomeFuncao", ""),
                "nomeSubFuncao": record.get("nomeSubFuncao", ""),
                "nomeAcaoOrcamentaria": record.get("nomeAcaoOrcamentaria", ""),
                **result,
            }
        )

    Path(args.log).parent.mkdir(parents=True, exist_ok=True)
    if rows or not Path(args.log).exists():
        write_table_rows(args.log, fields, rows, "log_classificacao")
    wb.save(args.workbook)
    summary = {
        "linhas_ja_preenchidas_antes": before,
        "linhas_aplicadas": len(rows),
        "indicador_nao_vazio_nas_linhas_aplicadas": indicator_nonempty_on_applied_rows,
        "distribuicao_area": dict(Counter(row["Area Tematica LLM"] for row in rows)),
        "distribuicao_gasto": dict(Counter(row["Gasto E ou NE LLM"] for row in rows)),
        "distribuicao_confianca": dict(Counter(row["Confianca"] for row in rows)),
        "distribuicao_revisao": dict(Counter(row["Requer revisao humana"] for row in rows)),
    }
    Path(args.summary).write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False))


def command_sample(args):
    random.seed(args.seed)
    if not Path(args.log).exists():
        return command_sample_workbook(args)
    rows = read_table_rows(args.log)
    required = {"excel_row", "Area Tematica LLM", "Gasto E ou NE LLM", "Requer revisao humana", "Confianca"}
    if not rows or not required.issubset(set(rows[0].keys())):
        return command_sample_workbook(args)
    if not rows:
        raise SystemExit("Log vazio; nada para amostrar.")
    review = [row for row in rows if row.get("Requer revisao humana") == "Sim"]
    high = [row for row in rows if row.get("Requer revisao humana") != "Sim" and row.get("Confianca") == "Alta"]
    groups = defaultdict(list)
    for row in high:
        groups[(row.get("nomeMunicipio", ""), row.get("Area Tematica LLM", ""))].append(row)
    sample = list(review)
    for group_rows in groups.values():
        sample.extend(random.sample(group_rows, min(args.per_group, len(group_rows))))
    fields = list(rows[0].keys()) + ["Revisao", "Area Tematica Revisada", "Gasto E ou NE Revisado", "Comentario Revisor"]
    seen = set()
    output_rows = []
    for row in sample:
        key = row.get("excel_row")
        if key in seen:
            continue
        seen.add(key)
        output_rows.append({**row, "Revisao": "", "Area Tematica Revisada": "", "Gasto E ou NE Revisado": "", "Comentario Revisor": ""})
    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    write_table_rows(args.output, fields, output_rows, "amostra_validacao")
    print(json.dumps({"rows": len(output_rows), "review_rows": len(review), "output": args.output}, ensure_ascii=False))


def command_sample_workbook(args):
    wb = load_workbook(args.workbook, read_only=True, data_only=True)
    ws = wb[args.sheet]
    headers = headers_for(ws)
    area_i = find_area_col(headers) - 1
    gasto_i = headers.index("Gasto E ou NE")
    wanted = ["Id", "nomeMunicipio", "nomePrograma", "nomeFuncao", "nomeSubFuncao", "nomeAcaoOrcamentaria"]
    col = {header: index for index, header in enumerate(headers)}
    groups = defaultdict(list)

    for row_num, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        area = clean(values[area_i])
        gasto = clean(values[gasto_i])
        if not area and not gasto:
            continue
        row = {"excel_row": row_num}
        for field in wanted:
            row[field] = clean(values[col[field]]) if field in col else ""
        row["Area Tematica LLM"] = area
        row["Gasto E ou NE LLM"] = gasto
        row["Confianca"] = ""
        row["Justificativa"] = "Amostra gerada diretamente da planilha classificada."
        row["Campos faltantes ou duvidas"] = ""
        row["Requer revisao humana"] = ""
        groups[(area, gasto)].append(row)

    sample = []
    for group_rows in groups.values():
        sample.extend(random.sample(group_rows, min(args.per_group, len(group_rows))))

    fields = ["excel_row", *wanted, "Area Tematica LLM", "Gasto E ou NE LLM", "Confianca", "Justificativa", "Campos faltantes ou duvidas", "Requer revisao humana", "Revisao", "Area Tematica Revisada", "Gasto E ou NE Revisado", "Comentario Revisor"]
    output_rows = []
    for row in sample:
        output_rows.append({**row, "Revisao": "", "Area Tematica Revisada": "", "Gasto E ou NE Revisado": "", "Comentario Revisor": ""})
    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    write_table_rows(args.output, fields, output_rows, "amostra_validacao")
    print(json.dumps({"rows": len(output_rows), "source": "workbook", "output": args.output}, ensure_ascii=False))


def command_apply_review(args):
    reviews = read_table_rows(args.review)
    wb = load_workbook(args.workbook)
    ws = wb[args.sheet]
    headers = headers_for(ws)
    area_col = find_area_col(headers)
    gasto_col = headers.index("Gasto E ou NE") + 1
    applied = 0
    doubts = 0
    ok = 0
    for row in reviews:
        status = norm(row.get("Revisao"))
        if status == "ok":
            ok += 1
            continue
        if status == "duvida":
            doubts += 1
            continue
        if status == "corrigir":
            excel_row = int(row["excel_row"])
            area = clean(row.get("Area Tematica Revisada"))
            gasto = clean(row.get("Gasto E ou NE Revisado"))
            if area:
                ws.cell(excel_row, area_col).value = area
            if gasto:
                ws.cell(excel_row, gasto_col).value = gasto
            applied += 1
    wb.save(args.workbook)
    print(json.dumps({"corrigidas": applied, "ok": ok, "duvidas": doubts}, ensure_ascii=False))


def command_report(args):
    status = workbook_status(args.workbook, args.sheet)
    wb = load_workbook(args.workbook, read_only=True, data_only=True)
    ws = wb[args.sheet]
    headers = headers_for(ws)
    area_i = find_area_col(headers) - 1
    gasto_i = headers.index("Gasto E ou NE")
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    log_rows = []
    if Path(args.log).exists():
        log_rows = read_table_rows(args.log)
    text = [
        "# Relatorio Final GSPI-M",
        "",
        f"Workbook: `{args.workbook}`",
        "",
        "## Totais",
        "",
        f"* Linhas de dados: {status['rows']}",
        f"* Area Tematica preenchida: {status['area_filled']}",
        f"* Gasto E ou NE preenchido: {status['gasto_filled']}",
        f"* Indicador preenchido: {status['indicador_filled']}",
        "",
        "## Distribuicao Por Area",
        "",
    ]
    for key, value in Counter(row[area_i] for row in rows).most_common():
        text.append(f"* {key}: {value}")
    text.extend(["", "## Distribuicao Por Tipo De Gasto", ""])
    for key, value in Counter(row[gasto_i] for row in rows).most_common():
        text.append(f"* {key}: {value}")
    if log_rows:
        text.extend(["", "## Revisao", "", f"* Linhas no log: {len(log_rows)}", f"* Marcadas para revisao: {sum(1 for row in log_rows if row.get('Requer revisao humana') == 'Sim')}"])
    Path(args.output).write_text("\n".join(text) + "\n", encoding="utf-8")
    print(json.dumps({"output": args.output}, ensure_ascii=False))


def command_run(args):
    source = Path(args.source)
    workbook = Path(args.workbook)
    workbook.parent.mkdir(parents=True, exist_ok=True)

    if args.force or not workbook.exists():
        shutil.copy2(source, workbook)
        prepared = "copied"
    else:
        prepared = "existing_workbook_used"

    classify_args = argparse.Namespace(
        workbook=str(workbook),
        sheet=args.sheet,
        log=args.log,
        summary=args.summary,
    )
    command_classify(classify_args)

    report_args = argparse.Namespace(
        workbook=str(workbook),
        sheet=args.sheet,
        log=args.log,
        output=args.report,
    )
    command_report(report_args)

    result = workbook_status(str(workbook), args.sheet)
    result.update(
        {
            "status": "ok",
            "source": str(source),
            "workbook": str(workbook),
            "workbook_preparation": prepared,
            "log": args.log,
            "summary": args.summary,
            "report": args.report,
        }
    )
    print(json.dumps(result, ensure_ascii=False))


def build_parser():
    parser = argparse.ArgumentParser(description="GSPI-M workflow CLI")
    sub = parser.add_subparsers(dest="command", required=True)

    p = sub.add_parser("prepare")
    p.add_argument("--source", default="fonte_parte_2.xlsx")
    p.add_argument("--output", default=str(DEFAULT_WORKBOOK))
    p.add_argument("--force", action="store_true")
    p.set_defaults(func=command_prepare)

    p = sub.add_parser("run")
    p.add_argument("--source", default="fonte_parte_2.xlsx")
    p.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    p.add_argument("--sheet", default=DEFAULT_SHEET)
    p.add_argument("--log", default=str(DEFAULT_LOG))
    p.add_argument("--summary", default=str(DEFAULT_SUMMARY))
    p.add_argument("--report", default="parte2/resultados/relatorio_final_classificacao.md")
    p.add_argument("--force", action="store_true")
    p.set_defaults(func=command_run)

    p = sub.add_parser("status")
    p.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    p.add_argument("--sheet", default=DEFAULT_SHEET)
    p.set_defaults(func=command_status)

    p = sub.add_parser("classify")
    p.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    p.add_argument("--sheet", default=DEFAULT_SHEET)
    p.add_argument("--log", default=str(DEFAULT_LOG))
    p.add_argument("--summary", default=str(DEFAULT_SUMMARY))
    p.set_defaults(func=command_classify)

    p = sub.add_parser("sample")
    p.add_argument("--log", default=str(DEFAULT_LOG))
    p.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    p.add_argument("--sheet", default=DEFAULT_SHEET)
    p.add_argument("--output", default="parte2/resultados/amostra_validacao_classificacao.xlsx")
    p.add_argument("--per-group", type=int, default=3)
    p.add_argument("--seed", type=int, default=42)
    p.set_defaults(func=command_sample)

    p = sub.add_parser("apply-review")
    p.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    p.add_argument("--sheet", default=DEFAULT_SHEET)
    p.add_argument("--review", required=True)
    p.set_defaults(func=command_apply_review)

    p = sub.add_parser("report")
    p.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    p.add_argument("--sheet", default=DEFAULT_SHEET)
    p.add_argument("--log", default=str(DEFAULT_LOG))
    p.add_argument("--output", default="parte2/resultados/relatorio_final_validacao.md")
    p.set_defaults(func=command_report)
    return parser


def main():
    args = build_parser().parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
