from __future__ import annotations

import argparse
import json
import unicodedata
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_SOURCE = Path("fonte_parte_2.xlsx")
DEFAULT_BASE = Path("parte2/municipios_base/municipios_base_expandido.xlsx")
DEFAULT_TEMPLATE = Path("parte2/modelos/modelo_municipios_base.xlsx")
BASE_SHEET = "municipios_classificados"
INPUT_SHEET = "novos_exemplos"

CANONICAL_HEADERS = [
    "Id",
    "codigoIbge",
    "nomeMunicipio",
    "nomeUnidadeGestora",
    "nomeOrgao",
    "nomeFuncao",
    "nomeSubFuncao",
    "nomeAcaoOrcamentaria",
    "nomePrograma",
    "grupoNaturezaDespesa",
    "idRubricaDespesa",
    "nomeRubricaDespesa",
    "valorEmpenhado",
    "valorLiquidado",
    "valorPago",
    "Area Tematica",
    "Gasto E ou NE",
    "Indicador",
    "Primeira Infancia",
    "Fonte Revisao",
    "Comentario Revisor",
]

REQUIRED_INPUT_HEADERS = [
    "Id",
    "nomeMunicipio",
    "nomePrograma",
    "nomeFuncao",
    "nomeSubFuncao",
    "nomeAcaoOrcamentaria",
    "Area Tematica",
    "Gasto E ou NE",
    "Indicador",
    "Fonte Revisao",
    "Comentario Revisor",
]

AREAS = {
    "Educacao Infantil",
    "Saude Materno-infantil",
    "Assistencia Social",
    "Protecao dos Direitos da Crianca e da Familia",
    "Direito a Cidade e Habitacao",
    "Saneamento e Agua",
    "Cultura e Direito de Brincar",
    "Seguranca Alimentar",
    "Enfrentamento da Pobreza",
    "nao se aplica",
}
GASTOS = {"Especifico", "Nao Especifico", "-"}


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def norm(value):
    text = unicodedata.normalize("NFKD", clean(value).lower())
    text = "".join(char for char in text if not unicodedata.combining(char))
    return " ".join(text.split())


def normalize_header(header):
    h = norm(header)
    if h in {"area tematica", "rea temtica"} or ("area" in h and "tem" in h):
        return "Area Tematica"
    if h in {"gasto e ou ne", "tipo de gasto"}:
        return "Gasto E ou NE"
    if h in {"primeira infancia", "primeira infncia"}:
        return "Primeira Infancia"
    for canonical in CANONICAL_HEADERS:
        if norm(canonical) == h:
            return canonical
    for canonical in REQUIRED_INPUT_HEADERS:
        if norm(canonical) == h:
            return canonical
    return clean(header)


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column_cells in ws.columns:
        values = [clean(cell.value) for cell in column_cells]
        width = min(max([len(value) for value in values] + [12]) + 2, 60)
        ws.column_dimensions[column_cells[0].column_letter].width = width
    ws.freeze_panes = "A2"


def read_sheet_rows(path, preferred_sheet=None):
    wb = load_workbook(path, read_only=True, data_only=True)
    if preferred_sheet and preferred_sheet in wb.sheetnames:
        ws = wb[preferred_sheet]
    elif INPUT_SHEET in wb.sheetnames:
        ws = wb[INPUT_SHEET]
    elif BASE_SHEET in wb.sheetnames:
        ws = wb[BASE_SHEET]
    else:
        ws = wb.active
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return [], []
    headers = [normalize_header(value) for value in raw_rows[0]]
    rows = []
    for values in raw_rows[1:]:
        row = {headers[index]: clean(value) for index, value in enumerate(values) if index < len(headers)}
        if any(row.values()):
            rows.append(row)
    return headers, rows


def write_workbook(path, sheet_name, headers, rows, instructions=None):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    style_sheet(ws)
    if instructions:
        wi = wb.create_sheet("instrucoes")
        for line in instructions:
            wi.append([line])
        wi.column_dimensions["A"].width = 110
    wb.save(path)


def row_key(row):
    if clean(row.get("Id")):
        return ("id", clean(row.get("Id")))
    return (
        "natural",
        norm(row.get("nomeMunicipio")),
        norm(row.get("nomePrograma")),
        norm(row.get("nomeFuncao")),
        norm(row.get("nomeSubFuncao")),
        norm(row.get("nomeAcaoOrcamentaria")),
    )


def canonicalize_row(row):
    result = {header: "" for header in CANONICAL_HEADERS}
    for key, value in row.items():
        normalized = normalize_header(key)
        if normalized in result:
            result[normalized] = clean(value)
    return result


def validate_rows(headers, rows):
    issues = []
    normalized_headers = {normalize_header(header) for header in headers}
    for header in REQUIRED_INPUT_HEADERS:
        if header not in normalized_headers:
            issues.append({"linha": 1, "tipo": "coluna ausente", "detalhe": header})
    seen = set()
    valid_rows = []
    for index, row in enumerate(rows, start=2):
        canonical = canonicalize_row(row)
        area = canonical["Area Tematica"]
        gasto = canonical["Gasto E ou NE"]
        if not canonical["nomeMunicipio"] or not canonical["nomeAcaoOrcamentaria"]:
            issues.append({"linha": index, "tipo": "linha incompleta", "detalhe": "nomeMunicipio e nomeAcaoOrcamentaria sao obrigatorios"})
            continue
        if not area or not gasto:
            issues.append({"linha": index, "tipo": "classificacao ausente", "detalhe": "Area Tematica e Gasto E ou NE devem estar preenchidos"})
            continue
        if area not in AREAS:
            issues.append({"linha": index, "tipo": "area invalida", "detalhe": area})
            continue
        if gasto not in GASTOS:
            issues.append({"linha": index, "tipo": "gasto invalido", "detalhe": gasto})
            continue
        if area == "nao se aplica" and gasto != "-":
            issues.append({"linha": index, "tipo": "combinacao invalida", "detalhe": "nao se aplica deve usar Gasto E ou NE = -"})
            continue
        key = row_key(canonical)
        if key in seen:
            issues.append({"linha": index, "tipo": "duplicidade no arquivo", "detalhe": str(key)})
            continue
        seen.add(key)
        valid_rows.append(canonical)
    return valid_rows, issues


def load_existing_base(source, base):
    rows = []
    if Path(base).exists():
        _, rows = read_sheet_rows(base, BASE_SHEET)
    elif Path(source).exists():
        _, rows = read_sheet_rows(source, BASE_SHEET)
    return [canonicalize_row(row) for row in rows]


def write_validation_report(path, valid_rows, issues):
    headers = ["linha", "tipo", "detalhe"]
    write_workbook(path, "problemas", headers, issues, ["Corrija as linhas listadas antes de expandir a base."])
    if valid_rows:
        wb = load_workbook(path)
        ws = wb.create_sheet("linhas_validas")
        ws.append(CANONICAL_HEADERS)
        for row in valid_rows:
            ws.append([row.get(header, "") for header in CANONICAL_HEADERS])
        style_sheet(ws)
        wb.save(path)


def command_template(args):
    examples = [
        {
            "Id": "exemplo-1",
            "nomeMunicipio": "Municipio Exemplo",
            "nomePrograma": "Programa Primeira Infancia",
            "nomeFuncao": "EDUCACAO",
            "nomeSubFuncao": "EDUCACAO INFANTIL",
            "nomeAcaoOrcamentaria": "MANUTENCAO DE CRECHES",
            "Area Tematica": "Educacao Infantil",
            "Gasto E ou NE": "Especifico",
            "Indicador": "",
            "Fonte Revisao": "exemplo",
            "Comentario Revisor": "Linha ilustrativa. Substituir por exemplo real.",
        }
    ]
    instructions = [
        "Preencha apenas exemplos humanos revisados.",
        "Nao inclua linhas ainda duvidosas.",
        "Use os rotulos exatamente como aparecem nas listas de apoio.",
        "Area Tematica e Gasto E ou NE sao obrigatorios para entrar na base.",
        "Indicador pode ficar em branco se ainda estiver fora do escopo do projeto.",
    ]
    write_workbook(args.output, INPUT_SHEET, REQUIRED_INPUT_HEADERS, examples, instructions)
    wb = load_workbook(args.output)
    wa = wb.create_sheet("rotulos_aceitos")
    wa.append(["Area Tematica"])
    for area in sorted(AREAS):
        wa.append([area])
    wg = wb.create_sheet("tipos_gasto")
    wg.append(["Gasto E ou NE"])
    for gasto in ["Especifico", "Nao Especifico", "-"]:
        wg.append([gasto])
    style_sheet(wa)
    style_sheet(wg)
    wb.save(args.output)
    print(json.dumps({"output": args.output}, ensure_ascii=False))


def command_init(args):
    if not Path(args.source).exists():
        raise SystemExit(f"Arquivo fonte nao encontrado: {args.source}")
    _, rows = read_sheet_rows(args.source, BASE_SHEET)
    canonical_rows = [canonicalize_row(row) for row in rows]
    write_workbook(args.base, BASE_SHEET, CANONICAL_HEADERS, canonical_rows, ["Base inicial criada a partir da aba municipios_classificados."])
    print(json.dumps({"status": "ok", "source": args.source, "base": args.base, "rows": len(canonical_rows)}, ensure_ascii=False))


def command_validate(args):
    headers, rows = read_sheet_rows(args.input, args.sheet)
    valid_rows, issues = validate_rows(headers, rows)
    report = args.report or str(Path(args.input).with_name(Path(args.input).stem + "_validacao.xlsx"))
    write_validation_report(report, valid_rows, issues)
    print(json.dumps({"valid_rows": len(valid_rows), "issues": len(issues), "report": report}, ensure_ascii=False))


def command_expand(args):
    headers, rows = read_sheet_rows(args.input, args.sheet)
    valid_rows, issues = validate_rows(headers, rows)
    if issues and not args.allow_issues:
        report = args.report or str(Path(args.input).with_name(Path(args.input).stem + "_validacao.xlsx"))
        write_validation_report(report, valid_rows, issues)
        raise SystemExit(json.dumps({"status": "blocked", "reason": "corrija os problemas antes de expandir", "issues": len(issues), "report": report}, ensure_ascii=False))

    existing = load_existing_base(args.source, args.base)
    keys = {row_key(row) for row in existing}
    added = []
    skipped_duplicates = 0
    for row in valid_rows:
        key = row_key(row)
        if key in keys:
            skipped_duplicates += 1
            continue
        keys.add(key)
        added.append(row)
    combined = existing + added
    write_workbook(args.base, BASE_SHEET, CANONICAL_HEADERS, combined, ["Base expandida de exemplos humanos GSPI-M."])
    summary = {
        "status": "ok",
        "base": args.base,
        "existing_rows": len(existing),
        "valid_input_rows": len(valid_rows),
        "added_rows": len(added),
        "skipped_duplicates": skipped_duplicates,
        "total_rows": len(combined),
        "area_distribution": dict(Counter(row["Area Tematica"] for row in combined)),
    }
    print(json.dumps(summary, ensure_ascii=False))


def command_sync(args):
    if not Path(args.base).exists():
        raise SystemExit(f"Base expandida nao encontrada: {args.base}")
    _, base_rows = read_sheet_rows(args.base, BASE_SHEET)
    base_rows = [canonicalize_row(row) for row in base_rows]
    wb = load_workbook(args.workbook)
    if BASE_SHEET in wb.sheetnames:
        del wb[BASE_SHEET]
    ws = wb.create_sheet(BASE_SHEET, 0)
    ws.append(CANONICAL_HEADERS)
    for row in base_rows:
        ws.append([row.get(header, "") for header in CANONICAL_HEADERS])
    style_sheet(ws)
    wb.save(args.workbook)
    print(json.dumps({"status": "ok", "workbook": args.workbook, "rows_synced": len(base_rows)}, ensure_ascii=False))


def build_parser():
    parser = argparse.ArgumentParser(description="GSPI-M municipios-base manager")
    sub = parser.add_subparsers(dest="command", required=True)

    p = sub.add_parser("init")
    p.add_argument("--source", default=str(DEFAULT_SOURCE))
    p.add_argument("--base", default=str(DEFAULT_BASE))
    p.set_defaults(func=command_init)

    p = sub.add_parser("template")
    p.add_argument("--output", default=str(DEFAULT_TEMPLATE))
    p.set_defaults(func=command_template)

    p = sub.add_parser("validate")
    p.add_argument("--input", required=True)
    p.add_argument("--sheet", default=INPUT_SHEET)
    p.add_argument("--report")
    p.set_defaults(func=command_validate)

    p = sub.add_parser("expand")
    p.add_argument("--input", required=True)
    p.add_argument("--sheet", default=INPUT_SHEET)
    p.add_argument("--source", default=str(DEFAULT_SOURCE))
    p.add_argument("--base", default=str(DEFAULT_BASE))
    p.add_argument("--report")
    p.add_argument("--allow-issues", action="store_true")
    p.set_defaults(func=command_expand)

    p = sub.add_parser("sync")
    p.add_argument("--workbook", default=str(DEFAULT_SOURCE))
    p.add_argument("--base", default=str(DEFAULT_BASE))
    p.set_defaults(func=command_sync)

    return parser


def main():
    args = build_parser().parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
